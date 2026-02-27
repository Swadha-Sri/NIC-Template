from pathlib import Path
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from .models import District, SolarYearData


TOTAL_ROW_MARKERS = {"total", "grand total", "overall"}


def normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "")


def normalize_district_code(raw_code):
    if raw_code is None:
        return ""

    value = str(raw_code).strip()
    if not value:
        return ""

    try:
        return str(int(float(value)))
    except (TypeError, ValueError):
        return value


def parse_int(value):
    try:
        if value is None or value == "":
            return 0
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def expanded_year_label(year_label):
    if "-" not in year_label:
        return year_label

    start, end_suffix = year_label.split("-", 1)
    if len(start) != 4 or len(end_suffix) != 2 or not start.isdigit() or not end_suffix.isdigit():
        return year_label

    return f"{start}-{start[:2]}{end_suffix}"


def resolve_metric_column(header_map, metric_name, year_label):
    direct = metric_name
    if direct in header_map:
        return header_map[direct]

    metric_suffix = f"_{metric_name}"
    for normalized, index in header_map.items():
        if normalized.endswith(metric_suffix):
            return index

    year_token = year_label.replace("-", "")
    full_year_token = expanded_year_label(year_label).replace("-", "")
    for normalized, index in header_map.items():
        if metric_name in normalized and (year_token in normalized or full_year_token in normalized):
            return index

    return None


def parse_upload_rows(file_path, year_label):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    header_row = [normalize_header(value) for value in rows[0]]
    header_map = {header: idx for idx, header in enumerate(header_row) if header}

    district_code_idx = header_map.get("distcode")
    if district_code_idx is None:
        district_code_idx = header_map.get("districtcode")

    district_name_idx = header_map.get("district")
    if district_name_idx is None:
        district_name_idx = header_map.get("districtname")

    if district_code_idx is None or district_name_idx is None:
        raise ValidationError("Excel format is invalid. Distcode and District columns are required.")

    target_idx = resolve_metric_column(header_map, "target", year_label)
    booking_idx = resolve_metric_column(header_map, "booking", year_label)
    installed_idx = resolve_metric_column(header_map, "installed", year_label)
    rejected_idx = resolve_metric_column(header_map, "rejected", year_label)

    if None in (target_idx, booking_idx, installed_idx, rejected_idx):
        raise ValidationError("Excel format is invalid. Target, Booking, installed, and Rejected columns are required.")

    parsed_rows = []

    for row in rows[1:]:
        district_code = normalize_district_code(row[district_code_idx] if district_code_idx < len(row) else "")
        district_name = str(row[district_name_idx] if district_name_idx < len(row) else "").strip()

        if not district_code or not district_name:
            continue

        if district_name.lower() in TOTAL_ROW_MARKERS:
            continue

        parsed_rows.append(
            {
                "district_code": district_code,
                "district_name": district_name,
                "target": parse_int(row[target_idx] if target_idx < len(row) else 0),
                "booking": parse_int(row[booking_idx] if booking_idx < len(row) else 0),
                "installed": parse_int(row[installed_idx] if installed_idx < len(row) else 0),
                "rejected": parse_int(row[rejected_idx] if rejected_idx < len(row) else 0),
            }
        )

    return sorted(parsed_rows, key=lambda item: item["district_name"].lower())


def import_solar_data_from_upload(upload):
    if not upload.file:
        raise ValidationError("Uploaded file is missing.")

    if SolarYearData.objects.filter(year_label=upload.year_label).exists():
        raise ValidationError(
            f"Data for {upload.year_label} already exists. Delete that upload before importing again."
        )

    file_name = upload.file.name
    storage = upload.file.storage

    if not storage.exists(file_name):
        raise ValidationError("Uploaded file not found in storage.")

    parsed_rows = parse_upload_rows(Path(storage.path(file_name)), upload.year_label)

    if not parsed_rows:
        raise ValidationError("No district rows found to import in this file.")

    with transaction.atomic():
        batch = []
        for item in parsed_rows:
            district, _ = District.objects.get_or_create(
                code=item["district_code"], defaults={"name": item["district_name"]}
            )
            if district.name != item["district_name"]:
                district.name = item["district_name"]
                district.save(update_fields=["name", "updated_at"])

            batch.append(
                SolarYearData(
                    district=district,
                    year_label=upload.year_label,
                    target=item["target"],
                    booking=item["booking"],
                    installed=item["installed"],
                    rejected=item["rejected"],
                )
            )

        SolarYearData.objects.bulk_create(batch)

    return len(batch)


def delete_year_data(year_label):
    deleted_count, _ = SolarYearData.objects.filter(year_label=year_label).delete()
    return deleted_count
