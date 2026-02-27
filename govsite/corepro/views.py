import json
from collections import defaultdict
from pathlib import Path
from statistics import StatisticsError
from math import sqrt
from scipy import stats
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.http import FileResponse, Http404
from django.core.exceptions import ValidationError
from django.views.decorators.http import require_POST
from accounts.forms import LoginCaptchaForm
from openpyxl import load_workbook
from .dataloader import load_solar_excel
from .analytics import calculate_descriptive_metrics
from .forecasting import calculate_district_forecast
from .models import SolarDataUpload
from .solar_data_service import import_solar_data_from_upload, delete_year_data

# Create your views here.
def index(request):
    return render(request, 'public/index.html', {
        "captcha_form": LoginCaptchaForm()
    })

def inner(request):
    return render(request, 'public/inner.html')

def sitemap(request):
    return render(request, 'public/site-map.html')

def contactus(request):
    return render(request, 'public/contactus.html')

@login_required(login_url='login')
def dashboard(request):
    if request.method == "POST":
        if request.POST.get("confirm_upload") != "yes":
            messages.error(request, "Upload cancelled. Please confirm before uploading.")
            return redirect("dashboard")

        uploaded_file = request.FILES.get("solar_data_file")

        if not uploaded_file:
            messages.error(request, "Please choose a file to upload.")
            return redirect("dashboard")

        filename = Path(uploaded_file.name).name

        try:
            SolarDataUpload.validate_filename(filename)
        except ValidationError as exc:
            messages.error(request, exc.message)
            return redirect("dashboard")

        year_label = SolarDataUpload.extract_year_label(filename)
        if SolarDataUpload.objects.filter(year_label=year_label).exists():
            messages.error(
                request,
                f"A file for {year_label} is already uploaded. Please delete it before uploading again.",
            )
            return redirect("dashboard")

        try:
            upload = SolarDataUpload.objects.create(
                file=uploaded_file,
                original_filename=filename,
                year_label=year_label,
                uploaded_by=request.user,
            )

            imported_rows = import_solar_data_from_upload(upload)
            messages.success(
                request,
                f"{filename} uploaded successfully and {imported_rows} district rows were stored.",
            )
        except ValidationError as exc:
            if "upload" in locals() and upload.pk:
                if upload.file:
                    upload.file.delete(save=False)
                upload.delete()
            messages.error(request, exc.message)

        return redirect("dashboard")

    data = load_solar_excel()
    descriptive_results = calculate_descriptive_metrics(data)

    total_target = sum(row["target"] for row in data)
    total_booking = sum(row["booking"] for row in data)
    total_installed = sum(row["installed"] for row in data)
    total_rejected = sum(row["rejected"] for row in data)

    install_rate = (total_installed / total_target * 100) if total_target else 0
    reject_rate = (total_rejected / total_booking * 100) if total_booking else 0

    year_data = defaultdict(int)
    for row in data:
        year_data[row["year"]] += row["installed"]

    year_labels = sorted(year_data.keys())
    year_values = [year_data[year] for year in year_labels]

    district_totals = defaultdict(int)
    for row in data:
        code = str(row["Distcode"])
        district_totals[code] += row["installed"]

    sorted_districts = sorted(
        district_totals.items(),
        key=lambda x: x[1],
        reverse=True
    )

    top_10 = sorted_districts[:10]
    top_labels = [d[0] for d in top_10]
    top_installed = [d[1] for d in top_10]
    top_installed_percentages = [
        round((val / total_installed) * 100, 2) if total_installed else 0
        for val in top_installed
    ]

    efficiency_sorted = sorted(
        descriptive_results,
        key=lambda x: x["avg_efficiency"],
        reverse=True
    )[:10]

    eff_labels = [d["district"] for d in efficiency_sorted]
    eff_values = [round(d["avg_efficiency"] * 100, 2) for d in efficiency_sorted]

    bottom_10 = sorted(descriptive_results, key=lambda x: x["score"])[:10]
    bottom_labels = [d["district"] for d in bottom_10]
    bottom_scores = [d["score"] for d in bottom_10]

    uploaded_files = SolarDataUpload.objects.all()

    context = {
        "solar_total_target": f"{total_target:,}",
        "solar_total_booking": f"{total_booking:,}",
        "solar_total_installed": f"{total_installed:,}",
        "solar_reject_rate": f"{reject_rate:.1f}%",
        "solar_install_rate": f"{install_rate:.1f}%",
        "solar_year_labels": json.dumps(year_labels),
        "solar_year_installed": json.dumps(year_values),
        "solar_top_labels": json.dumps(top_labels),
        "solar_top_installed_percentage": json.dumps(top_installed_percentages),
        "solar_top_installed": json.dumps(top_installed),
        "solar_eff_labels": json.dumps(eff_labels),
        "solar_eff_values": json.dumps(eff_values),
        "bottom_labels": json.dumps(bottom_labels),
        "bottom_scores": json.dumps(bottom_scores),
        "descriptive_results": descriptive_results,
        "uploaded_files": uploaded_files,
    }

    return render(request, 'dashboard/dashboard.html', context)


@login_required(login_url='login')
def download_solar_pump_data(request):
    file_path = Path(__file__).resolve().parent / "data" / "SolarPumpData_2020-21.xlsx"

    if not file_path.exists():
        raise Http404("Requested file was not found.")

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=file_path.name,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@login_required(login_url='login')
def predictive_dashboard(request):
    from statistics import mean, median, mode, stdev, StatisticsError

    data = load_solar_excel()
    selected_year = request.GET.get("year", "all")
    raw_selected_districts = [str(d).strip() for d in request.GET.getlist("district") if str(d).strip()]
    explicit_none_selected = "__none__" in raw_selected_districts
    selected_districts = [district for district in raw_selected_districts if district != "__none__"]

    all_years = sorted({str(row["year"]).strip() for row in data if str(row.get("year", "")).strip()})
    all_districts = sorted({str(row["district"]).strip() for row in data if str(row.get("district", "")).strip()})

    if not selected_districts and not explicit_none_selected:
        selected_districts = all_districts[:]

    selected_district_set = set(selected_districts)

    districts_with_selected_year = {
        row["district"]
        for row in data
        if selected_year == "all" or row["year"] == selected_year
    }

    scoped_data = [
        row for row in data
        if row["district"] in districts_with_selected_year
        and row["district"] in selected_district_set
    ]

    district_history_totals = defaultdict(lambda: {"booking": 0, "rejected": 0})
    district_codes = {}
    for row in scoped_data:
        district = row.get("district")
        district_history_totals[district]["booking"] += int(row.get("booking", 0) or 0)
        district_history_totals[district]["rejected"] += int(row.get("rejected", 0) or 0)
        if district not in district_codes:
            district_codes[district] = str(row.get("Distcode", "") or "").strip()

    forecast_results = calculate_district_forecast(scoped_data)
    forecast_results = sorted(forecast_results, key=lambda x: x["district"].lower())

    for row in forecast_results:
        growth_value = float(row.get("growth_rate", 0) or 0)
        row["growth_rate_abs"] = round(abs(growth_value), 2)
        row["is_negative_growth"] = growth_value < 0
        row["distcode"] = district_codes.get(row["district"], "-") or "-"

        history = district_history_totals.get(row["district"], {"booking": 0, "rejected": 0})
        total_booking = history["booking"]
        total_rejected = history["rejected"]
        row["loss_handling_scope_rate"] = round((total_rejected / total_booking) * 100, 2) if total_booking else 0

    def safe_stat(data_points, func):
        try:
            if len(data_points) == 0:
                return "-"
            result = func(data_points)
            return round(result, 2) if isinstance(result, float) else result
        except (ValueError, StatisticsError):
            return "-"

    growth_rates = [float(row.get("growth_rate_abs", 0) or 0) for row in forecast_results]
    next_year_targets = [int(row.get("predicted_next_year_installed", 0) or 0) for row in forecast_results]
    loss_scope_rates = [float(row.get("loss_handling_scope_rate", 0) or 0) for row in forecast_results]

    chart_labels = [row["district"] for row in forecast_results]
    chart_values = [row["predicted_next_year_installed"] for row in forecast_results]

    context = {
        "selected_year": selected_year,
        "selected_districts": selected_districts,
        "all_years": all_years,
        "all_districts": all_districts,
        "forecast_results": forecast_results,
        "total_growth_rate": round(sum(growth_rates), 2),
        "total_next_year_target": sum(next_year_targets),
        "total_loss_scope_rate": round(sum(loss_scope_rates), 2),
        "mean_growth_rate": safe_stat(growth_rates, mean),
        "mean_next_year_target": safe_stat(next_year_targets, mean),
        "mean_loss_scope_rate": safe_stat(loss_scope_rates, mean),
        "median_growth_rate": safe_stat(growth_rates, median),
        "median_next_year_target": safe_stat(next_year_targets, median),
        "median_loss_scope_rate": safe_stat(loss_scope_rates, median),
        "mode_growth_rate": safe_stat(growth_rates, mode),
        "mode_next_year_target": safe_stat(next_year_targets, mode),
        "mode_loss_scope_rate": safe_stat(loss_scope_rates, mode),
        "std_dev_growth_rate": safe_stat(growth_rates, stdev),
        "std_dev_next_year_target": safe_stat(next_year_targets, stdev),
        "std_dev_loss_scope_rate": safe_stat(loss_scope_rates, stdev),
        "chart_labels": json.dumps(chart_labels),
        "chart_values": json.dumps(chart_values),
    }

    return render(request, "dashboard/predictive.html", context)


@login_required(login_url='login')
def advanced_ana_dashboard(request):
    data = load_solar_excel()

    all_years = sorted({str(row.get("year", "")).strip() for row in data if str(row.get("year", "")).strip()})
    selected_years = [str(year).strip() for year in request.GET.getlist("year") if str(year).strip()]

    if not selected_years:
        selected_years = all_years[:]

    selected_year_set = set(selected_years)

    filtered_data = [
        row
        for row in data
        if str(row.get("year", "")).strip() in selected_year_set
    ]

    district_totals = defaultdict(lambda: {
        "district": "",
        "distcode": "",
        "target": 0,
        "booking": 0,
        "installed": 0,
        "rejected": 0,
    })

    for row in filtered_data:
        district = str(row.get("district", "")).strip()
        if not district:
            continue
        district_row = district_totals[district]
        district_row["district"] = district
        if not district_row["distcode"]:
            district_row["distcode"] = str(row.get("Distcode", "") or "").strip()
        district_row["target"] += int(row.get("target", 0) or 0)
        district_row["booking"] += int(row.get("booking", 0) or 0)
        district_row["installed"] += int(row.get("installed", 0) or 0)
        district_row["rejected"] += int(row.get("rejected", 0) or 0)

    district_table_rows = sorted(
        district_totals.values(),
        key=lambda x: x["district"].lower()
    )

    def strength_label(r_value):
        magnitude = abs(r_value)
        if magnitude >= 0.8:
            return "Very Strong"
        if magnitude >= 0.6:
            return "Strong"
        if magnitude >= 0.4:
            return "Moderate"
        if magnitude >= 0.2:
            return "Weak"
        return "Very Weak"

    def build_correlation(x_key, y_key, title, purpose):
        points = []
        x_values = []
        y_values = []

        for row in district_table_rows:
            x_val = float(row.get(x_key, 0) or 0)
            y_val = float(row.get(y_key, 0) or 0)
            x_values.append(x_val)
            y_values.append(y_val)
            points.append({
                "district": row.get("district", "-"),
                "distcode": row.get("distcode", "-"),
                "x": x_val,
                "y": y_val,
            })

        n = len(points)
        x_unique = len(set(x_values))
        y_unique = len(set(y_values))

        if n < 3 or x_unique < 2 or y_unique < 2:
            return {
                "title": title,
                "purpose": purpose,
                "x_key": x_key,
                "y_key": y_key,
                "sample_size": n,
                "r": None,
                "p": None,
                "is_significant": False,
                "strength": "Not Available",
                "direction": "Not Available",
                "interpretation": "Insufficient variation to compute reliable correlation.",
                "points": points,
                "regression_line": [],
            }

        r_value, p_value = stats.pearsonr(x_values, y_values)
        slope, intercept, _, _, _ = stats.linregress(x_values, y_values)
        r_value = float(r_value)
        p_value = float(p_value)
        slope = float(slope)
        intercept = float(intercept)
        r_squared = r_value ** 2
        min_x = min(x_values)
        max_x = max(x_values)
        regression_line = [
            {"x": float(min_x), "y": float(slope * min_x + intercept)},
            {"x": float(max_x), "y": float(slope * max_x + intercept)},
        ]

        # Improvement 1: Use ±0.05 threshold for neutral case
        if abs(r_value) < 0.05:
            direction = "No Linear Relationship"
        elif r_value > 0:
            direction = "Positive"
        else:
            direction = "Negative"

        # Improvement 2: Format p-values professionally
        p_display = "< 0.001" if p_value < 0.001 else round(p_value, 4)
        
        significance_text = "statistically significant" if p_value < 0.05 else "not statistically significant"
        # Improvement 3: Add R² to interpretation
        interpretation = f"{strength_label(r_value)} {direction} association; {significance_text} (p = {p_display}). R² = {round(r_squared, 4)} ({round(r_squared * 100, 2)}% variation explained)."

        return {
            "title": title,
            "purpose": purpose,
            "x_key": x_key,
            "y_key": y_key,
            "sample_size": n,
            "r": round(float(r_value), 4),
            "p": round(float(p_value), 6),
            "p_display": p_display,
            "r_squared": round(float(r_squared), 4),
            "r_squared_percent": round(float(r_squared) * 100, 2),
            "is_significant": bool(p_value < 0.05),
            "strength": strength_label(r_value),
            "direction": direction,
            "interpretation": interpretation,
            "points": points,
            "regression_line": regression_line,
        }

    correlation_results = [
        build_correlation(
            "booking",
            "installed",
            "Booking ↔ Installed",
            "Measures structural alignment of pipeline volumes across districts."
        ),
        build_correlation(
            "rejected",
            "installed",
            "Rejected ↔ Installed",
            "Tests whether higher rejection levels are associated with lower installation outcomes."
        ),
        build_correlation(
            "target",
            "installed",
            "Target ↔ Installed",
            "Checks planning alignment between assigned target and realized installation."
        ),
    ]

    context = {
        "all_years": all_years,
        "selected_years": selected_years,
        "district_count": len(district_table_rows),
        "correlation_results": correlation_results,
        "correlation_results_json": json.dumps(correlation_results),
    }

    return render(request, "dashboard/advanced_ana.html", context)


@login_required(login_url='login')
def descriptive_dashboard(request):
    from statistics import mean, median, mode, stdev
    
    data = load_solar_excel()

    all_years = sorted({str(row["year"]).strip() for row in data if str(row.get("year", "")).strip()})
    selected_years = [str(year).strip() for year in request.GET.getlist("year") if str(year).strip()]

    if not selected_years:
        selected_years = all_years[:]

    selected_year_set = set(selected_years)

    filtered_data = [
        row
        for row in data
        if str(row.get("year", "")).strip() in selected_year_set
    ]

    total_target = sum(row["target"] for row in filtered_data)
    total_booking = sum(row["booking"] for row in filtered_data)
    total_installed = sum(row["installed"] for row in filtered_data)
    total_rejected = sum(row["rejected"] for row in filtered_data)

    install_rate = (total_installed / total_target * 100) if total_target else 0
    reject_rate = (total_rejected / total_booking * 100) if total_booking else 0

    district_totals = defaultdict(lambda: {
        "district": "",
        "distcode": "",
        "target": 0,
        "booking": 0,
        "installed": 0,
        "rejected": 0,
    })
    district_installed_series = defaultdict(list)

    for row in filtered_data:
        district = row["district"]
        district_row = district_totals[district]
        district_row["district"] = district
        if not district_row["distcode"]:
            district_row["distcode"] = str(row.get("Distcode", "") or "").strip()
        district_row["target"] += int(row.get("target", 0) or 0)
        district_row["booking"] += int(row.get("booking", 0) or 0)
        district_row["installed"] += int(row.get("installed", 0) or 0)
        district_row["rejected"] += int(row.get("rejected", 0) or 0)
        district_installed_series[district].append(int(row.get("installed", 0) or 0))

    district_table_rows = []
    for district_row in district_totals.values():
        target = district_row["target"]
        booking = district_row["booking"]
        installed = district_row["installed"]
        rejected = district_row["rejected"]
        booking_difference = booking - installed
        
        tar = round((installed / target) * 100, 2) if target else 0
        brr = round((booking / target) * 100, 2) if target else 0
        icr = round((installed / booking) * 100, 2) if booking else 0
        rr = round((rejected / booking) * 100, 2) if booking else 0

        district_table_rows.append({
            "district": district_row["district"],
            "distcode": district_row["distcode"],
            "target": target,
            "booking": booking,
            "installed": installed,
            "rejected": rejected,
            "target_achievement_rate": tar,
            "booking_response_rate": brr,
            "installation_conversion_rate": icr,
            "rejection_rate": rr,
            "booking_difference": booking_difference,
        })

    top_installed_rows = sorted(district_table_rows, key=lambda x: x["installed"], reverse=True)[:10]
    bottom_installed_rows = sorted(district_table_rows, key=lambda x: x["installed"])[:10]
    top_rejected_rows = sorted(district_table_rows, key=lambda x: x["rejected"], reverse=True)[:10]
    top_booking_diff_rows = sorted(district_table_rows, key=lambda x: x["booking_difference"], reverse=True)[:10]
    top_icr_rows = sorted(district_table_rows, key=lambda x: x["installation_conversion_rate"], reverse=True)[:10]
    district_table_rows = sorted(district_table_rows, key=lambda x: x["district"].lower())

    # Calculate statistics
    targets = [row["target"] for row in district_table_rows]
    bookings = [row["booking"] for row in district_table_rows]
    installeds = [row["installed"] for row in district_table_rows]
    rejecteds = [row["rejected"] for row in district_table_rows]
    tars = [row["target_achievement_rate"] for row in district_table_rows]
    brrs = [row["booking_response_rate"] for row in district_table_rows]
    icrs = [row["installation_conversion_rate"] for row in district_table_rows]
    rrs = [row["rejection_rate"] for row in district_table_rows]
    booking_diffs = [row["booking_difference"] for row in district_table_rows]

    def safe_stat(data, func):
        try:
            if len(data) == 0:
                return "-"
            result = func(data)
            return round(result, 2) if isinstance(result, float) else result
        except (ValueError, StatisticsError):
            return "-"

    def safe_numeric_stat(data, func):
        try:
            if len(data) == 0:
                return 0
            return round(float(func(data)), 2)
        except (ValueError, StatisticsError, TypeError):
            return 0

    def coefficient_of_variation(data):
        """Calculate relative variability (CV = std dev / mean)"""
        try:
            if len(data) < 2:
                return "-"
            m = mean(data)
            if m == 0:
                return "-"
            return round((stdev(data) / m) * 100, 2)
        except (ValueError, StatisticsError):
            return "-"

    def calculate_skewness(data):
        """Calculate skewness"""
        try:
            if len(data) < 3:
                return "-"
            return round(stats.skew(data), 2)
        except:
            return "-"

    # Add relative variability and skewness to each district row
    for row in district_table_rows:
        district_series = district_installed_series.get(row["district"], [])
        row["relative_variability"] = coefficient_of_variation(district_series)
        row["skewness"] = calculate_skewness(district_series)
        skewness_value = row["skewness"]
        if isinstance(skewness_value, (int, float)):
            row["distribution_type"] = (
                "Right Skewed" if skewness_value > 0.5
                else "Left Skewed" if skewness_value < -0.5
                else "Approximately Symmetric"
            )
        else:
            row["distribution_type"] = "-"

    context = {
        "selected_years": selected_years,
        "all_years": all_years,
        "solar_total_target": f"{total_target:,}",
        "solar_total_booking": f"{total_booking:,}",
        "solar_total_installed": f"{total_installed:,}",
        "solar_total_rejected": f"{total_rejected:,}",
        "solar_reject_rate": f"{reject_rate:.1f}%",
        "solar_install_rate": f"{install_rate:.1f}%",
        "district_table_rows": district_table_rows,
        # Statistics
        "total_target": total_target,
        "total_booking": total_booking,
        "total_installed": total_installed,
        "total_rejected": total_rejected,
        "mean_target": safe_stat(targets, mean),
        "mean_booking": safe_stat(bookings, mean),
        "mean_installed": safe_stat(installeds, mean),
        "mean_rejected": safe_stat(rejecteds, mean),
        "mean_tar": safe_stat(tars, mean),
        "mean_brr": safe_stat(brrs, mean),
        "mean_icr": safe_stat(icrs, mean),
        "mean_rr": safe_stat(rrs, mean),
        "median_target": safe_stat(targets, median),
        "median_booking": safe_stat(bookings, median),
        "median_installed": safe_stat(installeds, median),
        "median_rejected": safe_stat(rejecteds, median),
        "median_tar": safe_stat(tars, median),
        "median_brr": safe_stat(brrs, median),
        "median_icr": safe_stat(icrs, median),
        "median_rr": safe_stat(rrs, median),
        "mode_target": safe_stat(targets, mode),
        "mode_booking": safe_stat(bookings, mode),
        "mode_installed": safe_stat(installeds, mode),
        "mode_rejected": safe_stat(rejecteds, mode),
        "mode_tar": safe_stat(tars, mode),
        "mode_brr": safe_stat(brrs, mode),
        "mode_icr": safe_stat(icrs, mode),
        "mode_rr": safe_stat(rrs, mode),
        "std_dev_target": safe_stat(targets, stdev),
        "std_dev_booking": safe_stat(bookings, stdev),
        "std_dev_installed": safe_stat(installeds, stdev),
        "std_dev_rejected": safe_stat(rejecteds, stdev),
        "std_dev_tar": safe_stat(tars, stdev),
        "std_dev_brr": safe_stat(brrs, stdev),
        "std_dev_icr": safe_stat(icrs, stdev),
        "std_dev_rr": safe_stat(rrs, stdev),
        # Relative variability (Coefficient of Variation)
        "cv_installed": coefficient_of_variation(installeds),
        "cv_icr": coefficient_of_variation(icrs),
        # Skewness
        "skew_installed": calculate_skewness(installeds),
        "skew_icr": calculate_skewness(icrs),
        # Chart data
        "solar_top_labels": json.dumps([row["district"] for row in top_installed_rows]),
        "solar_top_installed": json.dumps([row["installed"] for row in top_installed_rows]),
        "mean_installed_value": safe_numeric_stat(installeds, mean),
        "bottom_labels": json.dumps([row["district"] for row in bottom_installed_rows]),
        "bottom_installed": json.dumps([row["installed"] for row in bottom_installed_rows]),
        "rejected_diff_labels": json.dumps([row["district"] for row in top_rejected_rows]),
        "rejected_diff_values": json.dumps([row["rejected"] for row in top_rejected_rows]),
        "booking_diff_labels": json.dumps([row["district"] for row in top_booking_diff_rows]),
        "booking_diff_values": json.dumps([row["booking_difference"] for row in top_booking_diff_rows]),
        "icr_labels": json.dumps([row["district"] for row in top_icr_rows]),
        "icr_values": json.dumps([row["installation_conversion_rate"] for row in top_icr_rows]),
        "mean_icr_value": safe_numeric_stat(icrs, mean),
        "mean_rejected_value": safe_numeric_stat(rejecteds, mean),
        "mean_booking_diff_value": safe_numeric_stat(booking_diffs, mean),
    }

    return render(request, 'dashboard/descriptive.html', context)


@login_required(login_url='login')
def anamap_dashboard(request):
    data = load_solar_excel()

    year_totals_by_district = defaultdict(lambda: defaultdict(lambda: {
        "district": "",
        "target": 0,
        "installed": 0,
    }))
    average_totals_by_district = defaultdict(lambda: {
        "district": "",
        "target": 0,
        "installed": 0,
    })

    years = set()

    for row in data:
        district_code = str(row.get("Distcode", "")).strip()
        if not district_code:
            continue

        district_name = str(row.get("district", "")).strip()
        year = str(row.get("year", "")).strip()
        target = int(row.get("target", 0) or 0)
        installed = int(row.get("installed", 0) or 0)

        if not year:
            continue

        years.add(year)

        year_row = year_totals_by_district[year][district_code]
        year_row["district"] = district_name
        year_row["target"] += target
        year_row["installed"] += installed

        avg_row = average_totals_by_district[district_code]
        avg_row["district"] = district_name
        avg_row["target"] += target
        avg_row["installed"] += installed

    def build_rate_payload(district_totals):
        payload = {}
        for district_code, values in district_totals.items():
            target = values["target"]
            installed = values["installed"]
            achievement_rate = round((installed / target) * 100, 2) if target else 0

            payload[district_code] = {
                "district": values["district"],
                "target": target,
                "installed": installed,
                "achievement_rate": achievement_rate,
            }

        return payload

    average_map_data = build_rate_payload(average_totals_by_district)
    year_map_data = {
        year: build_rate_payload(district_data)
        for year, district_data in year_totals_by_district.items()
    }

    context = {
        "map_years": json.dumps(sorted(years)),
        "average_map_data": json.dumps(average_map_data),
        "year_map_data": json.dumps(year_map_data),
    }

    return render(request, 'dashboard/anamap.html', context)

@login_required(login_url='login')
def buttons(request):
    return render(request, 'dashboard/buttons.html')

@login_required(login_url='login')
def cards(request):
    return render(request, 'dashboard/cards.html')

@login_required(login_url='login')
def table(request):
    return render(request, 'dashboard/table.html')

@login_required(login_url='login')
def typography(request):
    return render(request, 'dashboard/typography.html')

@login_required(login_url='login')
def icons(request):
    return render(request, 'dashboard/icons.html')

@login_required(login_url='login')
def forms(request):
    return render(request, 'dashboard/forms.html')

@login_required(login_url='login')
def areachart(request):
    return render(request, 'dashboard/charts/areachart.html')

@login_required(login_url='login')
def scatterchart(request):
    return render(request, 'dashboard/charts/scatterchart.html')

@login_required(login_url='login')
def polarareachart(request):
    return render(request, 'dashboard/charts/polarareachart.html')

@login_required(login_url='login')
def linechart(request):
    return render(request, 'dashboard/charts/linechart.html')

@login_required(login_url='login')
def barchart(request):
    return render(request, 'dashboard/charts/barchart.html')

@login_required(login_url='login')
def doughnut_piechart(request):
    return render(request, 'dashboard/charts/doughnut_piechart.html')


@login_required(login_url='login')
def view_uploaded_file_data(request, upload_id):
    upload = get_object_or_404(SolarDataUpload, id=upload_id)

    if not upload.file or not upload.file.storage.exists(upload.file.name):
        raise Http404("Uploaded file not found.")

    workbook = load_workbook(upload.file.path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if rows:
        headers = ["" if value is None else str(value) for value in rows[0]]
        table_rows = [
            ["" if value is None else value for value in row]
            for row in rows[1:]
        ]
    else:
        headers = []
        table_rows = []

    context = {
        "upload": upload,
        "headers": headers,
        "table_rows": table_rows,
    }
    return render(request, "dashboard/upload_data_view.html", context)


@login_required(login_url='login')
def download_uploaded_file(request, upload_id):
    upload = get_object_or_404(SolarDataUpload, id=upload_id)

    if not upload.file or not upload.file.storage.exists(upload.file.name):
        raise Http404("Uploaded file not found.")

    return FileResponse(
        upload.file.open("rb"),
        as_attachment=True,
        filename=upload.original_filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@login_required(login_url='login')
@require_POST
def delete_uploaded_file(request, upload_id):
    upload = get_object_or_404(SolarDataUpload, id=upload_id)
    filename = upload.original_filename
    year_label = upload.year_label

    delete_year_data(year_label)

    if upload.file:
        upload.file.delete(save=False)
    upload.delete()

    messages.success(request, f"{filename} deleted successfully.")
    return redirect("dashboard")